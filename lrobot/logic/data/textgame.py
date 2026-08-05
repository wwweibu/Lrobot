"""文字博弈数据相关"""

import ast
import json
import math
import re
from datetime import datetime

from config import database_query, database_update, path

# 全局同一时间只允许一场文字博弈
LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
TEXTGAME_PATH = path / "storage/file/command/textgame"
# 计分式子:允许的 ast 节点,防止 eval 注入
SAFE_NODES = (
    ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant,
    ast.Add, ast.Sub, ast.Mult, ast.Div, ast.USub, ast.UAdd,
)


def textgame_options_parse(text):
    """从题面提取选项集

    取行首出现的大写字母,从 A 起取最长连续序列(A/AB/ABC...),
    这样题面里混入的杂散大写字母不会把选项集撑大;行首取不到时退回全文扫描。
    """
    heads = set(re.findall(r"^\s*([A-Z])[^A-Za-z]", text, re.MULTILINE))
    heads |= set(re.findall(r"^\s*([A-Z])\s*[一-鿿]", text, re.MULTILINE))
    if "A" not in heads:
        heads = set(re.findall(r"[A-Z]", text))
    options = ""
    for letter in LETTERS:
        if letter not in heads:
            break
        options += letter
    return options


def textgame_floor(value):
    """向下保留两位小数,官方规则

    先把浮点表示噪声抹掉再取整,否则 8+4.2 会得到 12.200000000000001,
    向下取整后凭空少一分钱。
    """
    return math.floor(round(float(value) * 100, 6)) / 100


def textgame_round2(value):
    """四舍五入到两位小数

    本模块把 round 当作题号的变量名用,内建 round() 会被遮蔽,故单独提供。
    """
    value = float(value) * 100
    return (math.floor(value + 0.5) if value >= 0 else math.ceil(value - 0.5)) / 100


def textgame_expr_eval(expr, base):
    """计算一个选项的式子,返回新的累计分

    +2 / -1 / +7*0.6 加减,作用于累计分
    *2 / /2       乘除,作用于累计分
    =5            直接指定累计分
    2             不带符号视同 +2
    """
    expr = str(expr).strip().replace("＋", "+").replace("－", "-").replace("×", "*").replace("÷", "/")
    expr = expr.replace("，", "").replace(" ", "")
    if not expr:
        return float(base)
    op = "+"
    if expr[0] in "+-*/=":
        op, expr = ("+" if expr[0] == "-" else expr[0]), (expr if expr[0] == "-" else expr[1:])
    if not expr:
        return float(base)

    tree = ast.parse(expr, mode="eval")
    for node in ast.walk(tree):
        if not isinstance(node, SAFE_NODES):
            raise ValueError(f"式子里有不支持的写法: {expr}")
        if isinstance(node, ast.Constant) and not isinstance(node.value, (int, float)):
            raise ValueError(f"式子里有不支持的写法: {expr}")
    value = float(eval(compile(tree, "<expr>", "eval"), {"__builtins__": {}}, {}))

    base = float(base)
    if op == "+":
        return textgame_floor(base + value)
    if op == "*":
        return textgame_floor(base * value)
    if op == "/":
        if not value:
            raise ValueError("式子里出现了除以 0")
        return textgame_floor(base / value)
    return textgame_floor(value)


async def textgame_get():
    """获取当前未结束的文字博弈"""
    rows = await database_query(
        "SELECT * FROM user_textgame WHERE status <> '已结束' ORDER BY id DESC LIMIT 1"
    )
    return rows[0] if rows else None


async def textgame_last():
    """获取最近一场文字博弈,含已结束"""
    rows = await database_query("SELECT * FROM user_textgame ORDER BY id DESC LIMIT 1")
    return rows[0] if rows else None


async def textgame_create(dm, platform, group_id):
    """新建一场文字博弈,返回 game_id"""
    return await database_update(
        "INSERT INTO user_textgame (dm, platform, group_id, status, round, start_time) "
        "VALUES (%s, %s, %s, '报名中', 0, %s)",
        (str(dm), platform, str(group_id), datetime.now()),
    )


async def textgame_edit(game_id, **fields):
    """更新一场文字博弈的字段"""
    if not fields:
        return
    sets = ", ".join(f"{k} = %s" for k in fields)
    await database_update(
        f"UPDATE user_textgame SET {sets} WHERE id = %s", (*fields.values(), game_id)
    )


async def textgame_finish(game_id):
    """结束一场文字博弈"""
    await database_update(
        "UPDATE user_textgame SET status = '已结束', signup_open = 0, end_time = %s WHERE id = %s",
        (datetime.now(), game_id),
    )


async def textgame_players(game_id):
    """获取全部玩家,按报名先后"""
    return await database_query(
        "SELECT * FROM user_textgame_player WHERE game_id = %s ORDER BY id", (game_id,)
    )


async def textgame_player_get(game_id, user):
    """获取单个玩家"""
    rows = await database_query(
        "SELECT * FROM user_textgame_player WHERE game_id = %s AND user = %s LIMIT 1",
        (game_id, str(user)),
    )
    return rows[0] if rows else None


async def textgame_player_add(game_id, user, name):
    """报名,返回 (是否新增, 当前人数)"""
    exist = await textgame_player_get(game_id, user)
    if not exist:
        await database_update(
            "INSERT INTO user_textgame_player (game_id, user, name, join_time) VALUES (%s, %s, %s, %s)",
            (game_id, str(user), name, datetime.now()),
        )
    total = await database_query(
        "SELECT COUNT(*) AS c FROM user_textgame_player WHERE game_id = %s", (game_id,)
    )
    return (not exist), total[0]["c"]


async def textgame_player_delete(game_id, user):
    """退出/被移出名单"""
    await database_update(
        "DELETE FROM user_textgame_player WHERE game_id = %s AND user = %s", (game_id, str(user))
    )


async def textgame_round_get(game_id, round=None):
    """获取指定题;不传则取当前最新一题"""
    if round is None:
        rows = await database_query(
            "SELECT * FROM user_textgame_round WHERE game_id = %s ORDER BY round DESC LIMIT 1",
            (game_id,),
        )
    else:
        rows = await database_query(
            "SELECT * FROM user_textgame_round WHERE game_id = %s AND round = %s LIMIT 1",
            (game_id, round),
        )
    return rows[0] if rows else None


async def textgame_rounds(game_id):
    """获取全部题,按题号"""
    return await database_query(
        "SELECT * FROM user_textgame_round WHERE game_id = %s ORDER BY round", (game_id,)
    )


async def textgame_round_open(game_id, round, options=None, title=None):
    """开一题,重开同一题会清空该题已收的票"""
    exist = await textgame_round_get(game_id, round)
    if exist:
        await database_update(
            "DELETE FROM user_textgame_vote WHERE game_id = %s AND round = %s", (game_id, round)
        )
        await database_update(
            "UPDATE user_textgame_round SET status = '收集中', options = %s, title = %s, dist = NULL, "
            "start_time = %s, end_time = NULL WHERE id = %s",
            (options, title, datetime.now(), exist["id"]),
        )
    else:
        await database_update(
            "INSERT INTO user_textgame_round (game_id, round, status, options, title, start_time) "
            "VALUES (%s, %s, '收集中', %s, %s, %s)",
            (game_id, round, options, title, datetime.now()),
        )
    await textgame_edit(game_id, round=round, status="进行中")


async def textgame_options_set(game_id, round, options):
    """主持人手动改本题选项集

    填单个字母(如 F)视作 A 到 F;填一串(如 ABDF)则按这一串,允许跳字母。
    """
    letters = sorted(set(re.findall(r"[A-Z]", str(options).upper())))
    if not letters:
        options = ""
    elif len(letters) == 1:
        options = LETTERS[: LETTERS.index(letters[0]) + 1]
    else:
        options = "".join(letters)
    await database_update(
        "UPDATE user_textgame_round SET options = %s WHERE game_id = %s AND round = %s",
        (options, game_id, round),
    )
    return options


async def textgame_votes(game_id, round):
    """获取某题全部票"""
    return await database_query(
        "SELECT * FROM user_textgame_vote WHERE game_id = %s AND round = %s", (game_id, round)
    )


async def textgame_vote_get(game_id, round, user):
    """获取某人某题的票"""
    rows = await database_query(
        "SELECT * FROM user_textgame_vote WHERE game_id = %s AND round = %s AND user = %s LIMIT 1",
        (game_id, round, str(user)),
    )
    return rows[0] if rows else None


async def textgame_vote_set(game_id, round, user, choice):
    """提交或修改答案,返回 (结果, 最终选项)

    结果为 '收到'(首次) / '已改'(用掉唯一一次修改机会) / '不可改'(已用过)
    """
    choice = choice.upper()
    now = datetime.now()
    vote = await textgame_vote_get(game_id, round, user)
    if not vote:
        await database_update(
            "INSERT INTO user_textgame_vote (game_id, round, user, choice, changed, created_at, updated_at) "
            "VALUES (%s, %s, %s, %s, 0, %s, %s)",
            (game_id, round, str(user), choice, now, now),
        )
        return "收到", choice
    if vote["changed"]:
        return "不可改", vote["choice"]
    if vote["choice"] == choice:
        return "已改", choice  # 改成同一个也消耗机会,与手动重复提交等价
    await database_update(
        "UPDATE user_textgame_vote SET choice = %s, changed = 1, updated_at = %s WHERE id = %s",
        (choice, now, vote["id"]),
    )
    return "已改", choice


async def textgame_dist(game_id, round):
    """统计某题分布,返回 (分布串, {选项: 人数}, 未交票玩家列表)

    分布串按官方格式,如 3A1B3C2D,无人选择的选项不出现;未交票者不计入分布
    """
    votes = await textgame_votes(game_id, round)
    counter = {}
    voted = set()
    for v in votes:
        if not v["choice"]:
            continue
        counter[v["choice"]] = counter.get(v["choice"], 0) + 1
        voted.add(v["user"])
    dist = "".join(f"{counter[c]}{c}" for c in sorted(counter))
    players = await textgame_players(game_id)
    absent = [p for p in players if p["user"] not in voted]
    return dist, counter, absent


async def textgame_round_close(game_id, round):
    """截止一题,写入分布,返回 (分布串, {选项: 人数}, 未交票玩家列表)"""
    dist, counter, absent = await textgame_dist(game_id, round)
    await database_update(
        "UPDATE user_textgame_round SET status = '已截止', dist = %s, end_time = %s "
        "WHERE game_id = %s AND round = %s",
        (dist, datetime.now(), game_id, round),
    )
    return dist, counter, absent


async def textgame_ops_get(game_id, round):
    """获取某题的计分式子 {选项: 式子}"""
    row = await textgame_round_get(game_id, round)
    if not row or not row["ops"]:
        return {}
    try:
        return json.loads(row["ops"])
    except Exception:
        return {}


async def textgame_ops_set(game_id, round, ops):
    """写入某题的计分式子,先逐条试算,有写错的就整体不落库"""
    errors = []
    for option, expr in (ops or {}).items():
        try:
            textgame_expr_eval(expr, 0)
        except Exception as e:
            errors.append(f"{option} 的式子有问题: {e}")
    if errors:
        raise ValueError("；".join(errors))
    await database_update(
        "UPDATE user_textgame_round SET ops = %s WHERE game_id = %s AND round = %s",
        (json.dumps(ops, ensure_ascii=False), game_id, round),
    )


async def textgame_total(game_id, user, before=None):
    """取某人截至 before(不含) 之前的累计分,before 为空则取最新"""
    if before is None:
        rows = await database_query(
            "SELECT total FROM user_textgame_vote WHERE game_id = %s AND user = %s "
            "AND total IS NOT NULL ORDER BY round DESC LIMIT 1",
            (game_id, str(user)),
        )
    else:
        rows = await database_query(
            "SELECT total FROM user_textgame_vote WHERE game_id = %s AND user = %s AND round < %s "
            "AND total IS NOT NULL ORDER BY round DESC LIMIT 1",
            (game_id, str(user), before),
        )
    return float(rows[0]["total"]) if rows else 0.0


async def textgame_choice_set(game_id, round, user, choice):
    """主持人在页面上改某人某题的选项,不消耗玩家的修改机会"""
    choice = (choice or "").upper().strip() or None
    now = datetime.now()
    vote = await textgame_vote_get(game_id, round, user)
    if vote:
        await database_update(
            "UPDATE user_textgame_vote SET choice = %s, updated_at = %s WHERE id = %s",
            (choice, now, vote["id"]),
        )
    else:
        await database_update(
            "INSERT INTO user_textgame_vote (game_id, round, user, choice, changed, created_at, updated_at) "
            "VALUES (%s, %s, %s, %s, 1, %s, %s)",
            (game_id, round, str(user), choice, now, now),
        )
    await textgame_dist_refresh(game_id, round)


async def textgame_dist_refresh(game_id, round):
    """重新统计并写回某题的分布"""
    dist, counter, absent = await textgame_dist(game_id, round)
    await database_update(
        "UPDATE user_textgame_round SET dist = %s WHERE game_id = %s AND round = %s",
        (dist, game_id, round),
    )
    return dist, counter, absent


async def textgame_score_clear(game_id, round, user):
    """取消手动覆盖,该格恢复按式子自动算"""
    vote = await textgame_vote_get(game_id, round, user)
    if vote:
        await database_update(
            "UPDATE user_textgame_vote SET manual = 0, score = NULL, updated_at = %s WHERE id = %s",
            (datetime.now(), vote["id"]),
        )
    await textgame_recalc(game_id, 1, user)


async def textgame_score_set(game_id, round, user, score):
    """主持人手动覆盖某人某题的得分,之后自动重算后续累计"""
    now = datetime.now()
    vote = await textgame_vote_get(game_id, round, user)
    if vote:
        await database_update(
            "UPDATE user_textgame_vote SET score = %s, manual = 1, updated_at = %s WHERE id = %s",
            (score, now, vote["id"]),
        )
    else:
        await database_update(
            "INSERT INTO user_textgame_vote (game_id, round, user, score, manual, created_at, updated_at) "
            "VALUES (%s, %s, %s, %s, 1, %s, %s)",
            (game_id, round, str(user), score, now, now),
        )
    await textgame_recalc(game_id, 1, user)


async def textgame_recalc(game_id, from_round=1, user=None):
    """按式子重算分数,默认整行从第 1 题算起

    每题的语义:新累计 = 式子作用于旧累计,本题得分 = 新累计 - 旧累计。
    标红(手动改过)的格子不参与公式重算,其分数一律按加法计入累计。
    未交票者按本场的 absent_as_a 开关决定是否视同选 A。
    传 user 则只重算这一个人的一行,改单个格子时用。
    """
    game = await database_query("SELECT * FROM user_textgame WHERE id = %s", (game_id,))
    if not game:
        return []
    absent_as_a = bool(game[0]["absent_as_a"])
    players = await textgame_players(game_id)
    if user is not None:
        players = [p for p in players if str(p["user"]) == str(user)]
    rounds = [r for r in await textgame_rounds(game_id) if r["round"] >= from_round]
    errors = []

    for player in players:
        user = player["user"]
        total = await textgame_total(game_id, user, before=from_round)
        for round_row in rounds:
            round = round_row["round"]
            try:
                ops = json.loads(round_row["ops"]) if round_row["ops"] else {}
            except Exception:
                ops = {}
            vote = await textgame_vote_get(game_id, round, user)
            if vote and vote["manual"]:
                score = float(vote["score"] or 0)
                total = textgame_floor(total + score)  # 手动分本身可能是任意小数
            else:
                choice = (vote or {}).get("choice")
                if not choice and absent_as_a:
                    choice = "A"
                expr = ops.get(choice) if choice else None
                if expr:
                    try:
                        new_total = textgame_expr_eval(expr, total)
                    except Exception as e:
                        errors.append(f"第{round}题 {choice}: {e}")
                        new_total = total
                else:
                    new_total = total
                score = textgame_round2(new_total - total)
                total = new_total
            if vote:
                await database_update(
                    "UPDATE user_textgame_vote SET score = %s, total = %s WHERE id = %s",
                    (score, total, vote["id"]),
                )
            else:
                await database_update(
                    "INSERT INTO user_textgame_vote (game_id, round, user, score, total, created_at, updated_at) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (game_id, round, str(user), score, total, datetime.now(), datetime.now()),
                )
    return sorted(set(errors))


async def textgame_board(game_id):
    """页面用的全量数据"""
    game = await database_query("SELECT * FROM user_textgame WHERE id = %s", (game_id,))
    if not game:
        return None
    game = game[0]
    players = await textgame_players(game_id)
    rounds = await textgame_rounds(game_id)
    votes = await database_query(
        "SELECT * FROM user_textgame_vote WHERE game_id = %s", (game_id,)
    )
    vote_map = {}
    for v in votes:
        vote_map.setdefault(str(v["round"]), {})[v["user"]] = {
            "choice": v["choice"],
            "changed": bool(v["changed"]),
            "score": None if v["score"] is None else float(v["score"]),
            "total": None if v["total"] is None else float(v["total"]),
            "manual": bool(v["manual"]),
        }
    round_list = []
    for r in rounds:
        try:
            ops = json.loads(r["ops"]) if r["ops"] else {}
        except Exception:
            ops = {}
        _, counter, absent = await textgame_dist(game_id, r["round"])
        round_list.append(
            {
                "round": r["round"],
                "status": r["status"],
                "options": r["options"] or "",
                "title": r["title"] or "",
                "dist": r["dist"] or "",
                "counter": counter,
                "absent": [p["user"] for p in absent],
                "ops": ops,
            }
        )
    return {
        "game": {
            "id": game["id"],
            "dm": game["dm"],
            "group_id": game["group_id"],
            "status": game["status"],
            "round": game["round"],
            "signup_open": bool(game["signup_open"]),
            "absent_as_a": bool(game["absent_as_a"]),
        },
        "players": [{"user": p["user"], "name": p["name"]} for p in players],
        "rounds": round_list,
        "votes": vote_map,
        "rank": await textgame_rank(game_id),
    }


async def textgame_excel(game_id):
    """导出官方格式的复盘 excel,返回文件路径"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    board = await textgame_board(game_id)
    if not board:
        return None
    players, rounds = board["players"], board["rounds"]
    total_round = len(rounds) or 1
    rank_map = {r["user"]: r["rank"] for r in board["rank"]}

    wb = Workbook()
    ws = wb.active
    ws.title = str(total_round)

    # 表头:玩家 | 1 | 1小结 | 2 | 2小结 | ... | 总分 | 排名
    ws.cell(1, 1, "玩家")
    for i in range(total_round):
        ws.cell(1, 2 + i * 2, rounds[i]["round"] if i < len(rounds) else i + 1)
        ws.cell(1, 3 + i * 2, "总分" if i == total_round - 1 else f"{rounds[i]['round']}小结")
    rank_col = 2 + total_round * 2
    ws.cell(1, rank_col, "排名")

    # 第2行:每题的计分式子
    ws.cell(2, 1, "操作")
    for i, r in enumerate(rounds):
        ops = r["ops"] or {}
        ws.cell(2, 2 + i * 2, "  ".join(f"{k}{v}" for k, v in sorted(ops.items())))

    # 第3行:所有人 -> 分布
    ws.cell(3, 1, "所有人")
    for i, r in enumerate(rounds):
        ws.cell(3, 2 + i * 2, "/")
        ws.cell(3, 3 + i * 2, r["dist"] or "")

    # 每位玩家两行:上行选项,下行得分,小结列用公式联动
    for n, player in enumerate(players):
        top = 4 + n * 2
        bottom = top + 1
        ws.cell(top, 1, player["name"] or player["user"])
        ws.merge_cells(start_row=top, start_column=1, end_row=bottom, end_column=1)
        for i, r in enumerate(rounds):
            choice_col, sum_col = 2 + i * 2, 3 + i * 2
            cell = board["votes"].get(str(r["round"]), {}).get(player["user"], {})
            ws.cell(top, choice_col, cell.get("choice") or "")
            ws.cell(top, sum_col, "/")
            ws.cell(bottom, choice_col, cell.get("score"))
            left = get_column_letter(sum_col - 2)
            here = get_column_letter(choice_col)
            ws.cell(
                bottom, sum_col,
                f"={here}{bottom}" if i == 0 else f"={left}{bottom}+{here}{bottom}",
            )
        ws.cell(top, rank_col, "/")
        ws.cell(bottom, rank_col, rank_map.get(player["user"]))

    ws.column_dimensions["A"].width = 16
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B4"

    TEXTGAME_PATH.mkdir(parents=True, exist_ok=True)
    file = TEXTGAME_PATH / f"文字博弈复盘_{game_id}.xlsx"
    wb.save(file)
    return file


async def textgame_rank(game_id):
    """总分排名,返回 [{user, name, total, rank}],并列同名次"""
    players = await textgame_players(game_id)
    result = []
    for p in players:
        result.append(
            {"user": p["user"], "name": p["name"], "total": await textgame_total(game_id, p["user"])}
        )
    result.sort(key=lambda x: x["total"], reverse=True)
    rank = 0
    last = None
    for i, r in enumerate(result, 1):
        if r["total"] != last:
            rank = i
            last = r["total"]
        r["rank"] = rank
    return result
